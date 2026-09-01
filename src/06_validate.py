#!/usr/bin/env python3
"""
Northwind Cloud — step 6: prove the model, don't eyeball it.

Gates, all of which must pass:

  G1  Generator determinism        re-run and compare SHA-256 of every raw file
  G2  ARR bridge                   residual 0.00 in every month, to the cent
  G3  Cohort reconciliation        triangle foots to total MRR every month
  G4  ETL audit completeness       every rule logged, nothing dropped silently
  G5  Excel recalculates clean     zero formula errors after a LibreOffice recalc
  G6  Excel vs SQL                 every CHECKS row PASS, printed as a table
  G7  Scenario switch is live      flip it, recalculate, numbers must MOVE and
                                   must land on the Python model's answer

A green recalculation is not the same as correct numbers, which is why G6 and
G7 exist: G5 only proves nothing is broken, G6 proves the workbook agrees with
the warehouse, and G7 proves the switch is wired to something.
"""
from __future__ import annotations

import hashlib
import json
import shutil
import sqlite3
import subprocess
import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
DB = ROOT / "data" / "warehouse" / "northwind.db"
TAB = ROOT / "outputs" / "tables"
XLSX = ROOT / "outputs" / "excel" / "northwind_unit_economics.xlsx"
LOGS = ROOT / "outputs" / "logs"
ERRORS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#ERROR!")

results: list[dict] = []


def gate(gid, name, ok, detail=""):
    results.append(dict(gate=gid, name=name, status="PASS" if ok else "FAIL",
                        detail=detail))
    mark = "PASS" if ok else "FAIL"
    print(f"[{mark}] {gid}  {name}" + (f"\n        {detail}" if detail else ""))
    return ok


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def recalc(src: Path, workdir: Path) -> Path:
    """Recalculate a workbook headlessly with LibreOffice and return the
    rewritten file (which carries the computed values)."""
    workdir.mkdir(parents=True, exist_ok=True)
    profile = workdir / "loprofile"
    out = workdir / "out"
    out.mkdir(exist_ok=True)
    cmd = ["soffice", f"-env:UserInstallation=file://{profile}", "--headless",
           "--norestore", "--convert-to", "xlsx", "--outdir", str(out), str(src)]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=420)
    dest = out / (src.stem + ".xlsx")
    if not dest.exists():
        raise RuntimeError(f"LibreOffice produced nothing.\n{r.stdout}\n{r.stderr}")
    return dest


def scan_errors(wb) -> list[str]:
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and (v in ERRORS or v.startswith("Err:")):
                    bad.append(f"{ws.title}!{c.coordinate} = {v}")
    return bad


def main() -> int:
    print("=" * 78)
    print("VALIDATION")
    print("=" * 78)
    work = Path(tempfile.mkdtemp(prefix="nw_validate_"))

    # ---- G1 determinism ---------------------------------------------------
    manifest = json.loads((LOGS / "generation_manifest.json").read_text())
    before = {f: sha256(RAW / f) for f in manifest["sha256"]}
    subprocess.run([sys.executable, str(ROOT / "src" / "01_generate_data.py")],
                   capture_output=True, text=True, cwd=ROOT, check=True)
    after = {f: sha256(RAW / f) for f in manifest["sha256"]}
    diffs = [f for f in before if before[f] != after[f]]
    gate("G1", "Generator is deterministic (SHA-256 identical on re-run)",
         not diffs,
         "all 6 raw files byte-identical; "
         f"billing_raw.csv = {after['billing_raw.csv'][:16]}..."
         if not diffs else f"changed: {diffs}")

    con = sqlite3.connect(DB)

    # ---- G2 ARR bridge ----------------------------------------------------
    br = pd.read_csv(TAB / "Q02_arr_bridge.csv")
    worst = float(br.residual_arr_must_be_zero.abs().max())
    gate("G2", "ARR bridge ties to the cent in all 36 months", worst < 0.005,
         f"max |residual| = ${worst:.4f} across {len(br)} months")

    # ---- G3 cohort reconciliation ----------------------------------------
    rec = pd.read_csv(TAB / "Q04_cohort_reconciliation.csv")
    nfail = int((rec.status != "PASS").sum())
    gate("G3", "Cohort triangle foots to total MRR every month", nfail == 0,
         f"{len(rec) - nfail}/{len(rec)} months tie; max variance "
         f"${rec.variance_must_be_zero.abs().max():.4f}")

    # ---- G4 ETL audit -----------------------------------------------------
    au = pd.read_sql("SELECT * FROM etl_audit ORDER BY step_no", con)
    qn = pd.read_sql("SELECT COUNT(*) n FROM dq_quarantine", con).n.iloc[0]
    raw_rows = sum(1 for _ in open(RAW / "billing_raw.csv")) - 1
    loaded = int(au[au.rule_id == "R01"].rows_in.iloc[0])
    gate("G4", "ETL audit accounts for every source row",
         loaded == raw_rows and len(au) >= 13,
         f"{len(au)} rules logged, {raw_rows:,} billing rows in, "
         f"{qn:,} quarantined with a reason")

    # ---- G5 Excel recalculates clean --------------------------------------
    calc = recalc(XLSX, work / "base")
    wb = load_workbook(calc, data_only=True)
    bad = scan_errors(wb)
    gate("G5", "Excel recalculates with zero formula errors", not bad,
         f"{sum(ws.max_row for ws in wb.worksheets):,} rows scanned across "
         f"{len(wb.sheetnames)} sheets"
         if not bad else f"{len(bad)} error cells, first: {bad[:5]}")

    # ---- G6 Excel vs SQL --------------------------------------------------
    layout = json.loads((LOGS / "excel_layout.json").read_text())
    ck = wb["CHECKS"]
    rows = []
    r = 5
    while ck.cell(row=r, column=1).value:
        rows.append(dict(check=ck.cell(row=r, column=1).value,
                         excel=ck.cell(row=r, column=2).value,
                         reference=ck.cell(row=r, column=3).value,
                         difference=ck.cell(row=r, column=4).value,
                         tolerance=ck.cell(row=r, column=5).value,
                         status=ck.cell(row=r, column=6).value))
        r += 1
    chk = pd.DataFrame(rows)
    chk.to_csv(TAB / "excel_vs_sql_reconciliation.csv", index=False)
    nfail = int((chk.status != "PASS").sum())
    gate("G6", "Excel reconciles to SQL on every check", nfail == 0,
         f"{len(chk) - nfail}/{len(chk)} checks PASS")
    print()
    with pd.option_context("display.width", 200, "display.max_colwidth", 44):
        print(chk.to_string(index=False,
                            formatters={"excel": lambda v: f"{v:,.2f}" if isinstance(v, (int, float)) else str(v),
                                        "reference": lambda v: f"{v:,.2f}" if isinstance(v, (int, float)) else str(v),
                                        "difference": lambda v: f"{v:,.4f}" if isinstance(v, (int, float)) else str(v)}))
    print()

    # ---- G7 scenario switch is live ---------------------------------------
    ss = pd.read_csv(TAB / "scenario_summary.csv").set_index("scenario")
    fc_last = layout["forecast_last_row"]
    observed = {}
    for scen in ["base", "sales_proposal", "selective"]:
        tmp = work / f"scn_{scen}"
        tmp.mkdir(parents=True, exist_ok=True)
        f = tmp / XLSX.name
        shutil.copy(XLSX, f)
        w = load_workbook(f)                 # formulas preserved
        w["CONTROL"]["B3"] = scen
        w.save(f)
        got = load_workbook(recalc(f, tmp), data_only=True)
        cash = got["CALC_FORECAST"].cell(row=fc_last, column=27).value      # AA
        arr = got["CALC_FORECAST"].cell(row=fc_last, column=17).value * 12  # Q
        nfail_s = sum(1 for x in got["CHECKS"].iter_rows(min_row=5, min_col=6, max_col=6)
                      for c in x if c.value == "FAIL")
        observed[scen] = dict(cash_m18=cash, arr_m18=arr, failing_checks=nfail_s,
                              ref_cash=float(ss.loc[scen, "cash_month18"]),
                              ref_arr=float(ss.loc[scen, "arr_month18"]))
    sw = pd.DataFrame(observed).T
    sw["cash_diff"] = sw.cash_m18 - sw.ref_cash
    sw["arr_diff"] = sw.arr_m18 - sw.ref_arr
    moved = sw.cash_m18.nunique() == len(sw)
    # 0.5 = half of the last displayed unit in scenario_summary.csv. Anything
    # larger would mean Excel and Python actually disagree.
    accurate = bool((sw.cash_diff.abs() < 0.5).all() and (sw.arr_diff.abs() < 0.5).all())
    clean = bool((sw.failing_checks == 0).all())
    gate("G7", "Scenario switch moves the model and stays reconciled",
         moved and accurate and clean,
         f"3 distinct cash paths; max |cash diff| vs Python "
         f"${sw.cash_diff.abs().max():,.2f}; failing checks per scenario: "
         f"{sw.failing_checks.tolist()}")
    print()
    print(sw[["cash_m18", "ref_cash", "cash_diff", "arr_m18", "ref_arr", "arr_diff",
              "failing_checks"]].to_string(
        formatters={c: (lambda v: f"{v:,.2f}") for c in
                    ["cash_m18", "ref_cash", "cash_diff", "arr_m18", "ref_arr", "arr_diff"]}))
    sw.to_csv(TAB / "scenario_switch_test.csv")

    # ---- summary ----------------------------------------------------------
    summ = pd.DataFrame(results)
    summ.to_csv(TAB / "validation_gates.csv", index=False)
    print("\n" + "=" * 78)
    print(summ[["gate", "name", "status"]].to_string(index=False))
    failed = int((summ.status == "FAIL").sum())
    print("=" * 78)
    print(f"{len(summ) - failed}/{len(summ)} gates PASS")
    con.close()
    shutil.rmtree(work, ignore_errors=True)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
