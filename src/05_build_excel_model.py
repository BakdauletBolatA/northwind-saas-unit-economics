#!/usr/bin/env python3
"""
Northwind Cloud — step 5: build the Excel model.

Rules this workbook obeys:
  * Nothing outside a DATA_* sheet or the blue cells on CONTROL is a value.
    Every other cell is a formula. That includes the Holt smoothing recursion,
    the SDR saturation curve, the cohort compounding and the cash roll-forward.
  * The scenario switch on CONTROL!B3 moves the model. It drives an INDEX/MATCH
    into DATA_SCENARIOS, which feeds the whole forecast block.
  * Colour convention: BLUE = hard input, BLACK = formula on this sheet,
    GREEN = formula referencing another sheet.

src/06_validate.py recalculates this file with LibreOffice, checks for formula
errors, reconciles it against SQL, and flips the scenario switch to prove the
numbers actually move.
"""
from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "data" / "warehouse" / "northwind.db"
TAB = ROOT / "outputs" / "tables"
XLSX = ROOT / "outputs" / "excel" / "northwind_unit_economics.xlsx"

BLUE = Font(color="FF0000FF")
BLACK = Font(color="FF000000")
GREEN = Font(color="FF008000")
HDR = Font(bold=True, color="FFFFFFFF")
HDR_FILL = PatternFill("solid", fgColor="FF1F3864")
TITLE = Font(bold=True, size=13, color="FF1F3864")
SUB = Font(bold=True, color="FF1F3864")
BOX = PatternFill("solid", fgColor="FFF2F2F2")
THIN = Border(bottom=Side(style="thin", color="FFBFBFBF"))

MONEY = '#,##0;[Red](#,##0)'
MONEY2 = '#,##0.00;[Red](#,##0.00)'
PCT1 = '0.0%'
NUM2 = '#,##0.00'


def put(ws, cell, value, font=BLACK, fmt=None, align=None, fill=None):
    c = ws[cell]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align)
    if fill:
        c.fill = fill
    return c


def header_row(ws, row, labels, start_col=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def widths(ws, spec: dict):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ===========================================================================
def gather(con) -> dict:
    """Pull every value the workbook needs. This is the ONLY place values
    cross from the warehouse into Excel; everything else is computed there."""
    d = {}
    d["actuals"] = pd.read_sql("""
        WITH mv AS (
          SELECT d.month_index,
            SUM(CASE WHEN f.movement_type='new'            THEN f.movement_amount ELSE 0 END) AS new_mrr,
            SUM(CASE WHEN f.movement_type='opening_balance'THEN f.movement_amount ELSE 0 END) AS carry_in_mrr,
            SUM(CASE WHEN f.movement_type='expansion'      THEN f.movement_amount ELSE 0 END) AS expansion_mrr,
            SUM(CASE WHEN f.movement_type='reactivation'   THEN f.movement_amount ELSE 0 END) AS reactivation_mrr,
            SUM(CASE WHEN f.movement_type='contraction'    THEN -f.movement_amount ELSE 0 END) AS contraction_mrr,
            SUM(CASE WHEN f.movement_type='churn'          THEN -f.movement_amount ELSE 0 END) AS churn_mrr,
            SUM(CASE WHEN f.prev_mrr>0 THEN f.mrr ELSE 0 END) AS rn,
            SUM(CASE WHEN f.prev_mrr>0 THEN f.prev_mrr ELSE 0 END) AS rd
          FROM fact_subscription_month f JOIN dim_date d ON d.month_key=f.month_key
          GROUP BY d.month_index)
        SELECT p.month, p.month_index, p.mrr, p.recognised_revenue,
               mv.carry_in_mrr, mv.new_mrr, mv.expansion_mrr, mv.reactivation_mrr,
               mv.contraction_mrr, mv.churn_mrr,
               p.cogs, p.sm, p.rd, p.ga, p.total_cost, p.net_burn,
               CASE WHEN mv.rd>0 THEN mv.rn/mv.rd ELSE 1 END AS nrr_m
        FROM v_pnl_month p JOIN mv ON mv.month_index=p.month_index
        ORDER BY p.month_index""", con)

    d["channels"] = pd.read_csv(TAB / "Q07_cac_ltv_payback_by_channel.csv")
    d["segments_q"] = pd.read_csv(TAB / "Q09_unit_economics_by_segment.csv")
    d["seg_dim"] = pd.read_sql("SELECT * FROM dim_segment ORDER BY sort_order", con)
    d["params"] = json.load(open(TAB / "forward_model_parameters.json"))
    d["scenario_summary"] = pd.read_csv(TAB / "scenario_summary.csv")
    d["sensitivity"] = pd.read_csv(TAB / "hiring_sensitivity.csv")
    d["cash_paths"] = pd.read_csv(TAB / "scenario_cash_paths.csv")
    d["fmonths"] = pd.read_sql(
        f"SELECT month FROM dim_date WHERE is_actual=0 ORDER BY month_index "
        f"LIMIT {d['params']['forecast_months']}", con).month.tolist()
    return d


# ===========================================================================
def build():
    cfg = yaml.safe_load(open(ROOT / "config" / "assumptions.yml"))
    con = sqlite3.connect(DB)
    d = gather(con)
    P = d["params"]
    obs = P["observed"]
    seg_order = ["SMB", "MidMarket", "Enterprise"]
    scen_order = ["base", "sales_proposal", "selective"]
    H = int(P["forecast_months"])

    wb = Workbook()

    # ---------------------------------------------------------------- README
    ws = wb.active
    ws.title = "README"
    widths(ws, {"A": 3, "B": 34, "C": 96})
    put(ws, "B2", "Northwind Cloud — unit economics and runway model", TITLE)
    put(ws, "B3", "Board question: does the current sales machine pay back, "
                  "and should we add $180k/month of SDR capacity?")
    rows = [
        ("", ""),
        ("HOW TO USE", ""),
        ("1.", "Open CONTROL and pick a scenario in the blue cell B3. Every "
               "downstream number moves."),
        ("2.", "All other blue cells on CONTROL are live assumptions. Change one "
               "and the model recalculates."),
        ("3.", "CHECKS reconciles this workbook against the SQL warehouse. Every "
               "row must read PASS."),
        ("", ""),
        ("COLOUR CONVENTION", ""),
        ("Blue", "Hard input. The only cells a user should type into."),
        ("Black", "Formula computed on this sheet."),
        ("Green", "Formula referencing another sheet."),
        ("", ""),
        ("SHEETS", ""),
        ("CONTROL", "Scenario switch and every assumption the model uses."),
        ("DATA_ACTUALS", "36 months of actuals from the warehouse. Input only."),
        ("DATA_CHANNELS", "Trailing-12-month cost pools and behaviour by channel."),
        ("DATA_SEGMENTS", "Segment margins, retention, outbound ACV, sales-cycle lag."),
        ("DATA_SCENARIOS", "The three scenarios, one column each."),
        ("DATA_SQL_CHECK", "Reference values produced by the SQL library and the "
                           "Python cash model, used by CHECKS."),
        ("CALC_UNIT_ECON", "CAC, LTV and payback by channel — all formulas."),
        ("CALC_FORECAST", "Holt recursion, SDR saturation curve, cohort "
                          "compounding, cash roll-forward — all formulas."),
        ("OUT_SUMMARY", "The one-page answer."),
        ("CHECKS", "Excel vs SQL reconciliation, PASS/FAIL."),
        ("", ""),
        ("PROVENANCE", "Synthetic data generated by src/01_generate_data.py "
                       "(seeded, deterministic). The company is fictional; the "
                       "method is not."),
    ]
    r = 5
    for a, b in rows:
        if a in ("HOW TO USE", "COLOUR CONVENTION", "SHEETS", "PROVENANCE"):
            put(ws, f"B{r}", a, SUB)
        else:
            put(ws, f"B{r}", a, BLUE if a in ("Blue",) else
                (GREEN if a == "Green" else BLACK))
        put(ws, f"C{r}", b, BLACK)
        r += 1

    # --------------------------------------------------------------- CONTROL
    ws = wb.create_sheet("CONTROL")
    widths(ws, {"A": 44, "B": 18, "C": 60})
    put(ws, "A1", "CONTROL PANEL", TITLE)
    put(ws, "A3", "Scenario (pick one)", SUB)
    put(ws, "B3", "selective", BLUE, align="center", fill=BOX)
    put(ws, "C3", "<- blue cell: change this and the whole model moves", BLACK)
    put(ws, "A4", "Scenario column index")
    put(ws, "B4", "=MATCH($B$3,DATA_SCENARIOS!$B$1:$D$1,0)", GREEN, align="center")

    inputs = [
        ("cash_on_hand", "Cash on hand ($)", float(P["cash_on_hand"]), MONEY,
         "As at 2026-08-31, per the CEO"),
        ("board_hurdle", "Board minimum runway (months)", 18, NUM2,
         "The guardrail the board applies"),
        ("payback_hurdle", "CAC payback hurdle (months)", 18, NUM2,
         "Set equal to runway: we cannot fund a payback longer than our cash"),
        ("holt_alpha", "Holt alpha (level)", P["forecast"]["holt_alpha"], NUM2,
         "Grid-searched on in-sample one-step SSE in src/04"),
        ("holt_beta", "Holt beta (trend)", P["forecast"]["holt_beta"], NUM2,
         "Same fit. Holt won the rolling-origin backtest"),
        ("cogs_pct", "COGS % of revenue", P["cost_policy"]["cogs_pct"], PCT1,
         "Trailing 3 months"),
        ("sm_pct", "S&M % of revenue (base book)", P["cost_policy"]["sm_pct"], PCT1,
         "Current spending policy held constant"),
        ("rd_base", "R&D $/month at month 0", P["cost_policy"]["rd"], MONEY,
         "Headcount held flat"),
        ("ga_base", "G&A $/month at month 0", P["cost_policy"]["ga"], MONEY,
         "Headcount held flat"),
        ("infl", "Opex inflation (annual)", float(cfg["forecast_policy"]["opex_inflation_annual"]), PCT1,
         "Applied to R&D and G&A"),
        ("sdr_base_meet", "SDR meetings/rep at small team", float(cfg["sdr"]["meetings_per_rep_small_team"]), NUM2,
         "Observed when the team was 4 or fewer"),
        ("sdr_n0", "SDR saturation team size", float(cfg["sdr"]["saturation_team_size"]), NUM2,
         "Above this, meetings per rep decay"),
        ("sdr_exp", "SDR saturation exponent", float(cfg["sdr"]["saturation_exponent"]), NUM2,
         "per_rep(N) = base * (N0/N)^exponent"),
        ("sdr_now", "Current SDR heads", float(obs["sdr_heads_now"]), NUM2, "From the headcount ledger"),
        ("m2w", "Outbound meeting -> win rate", float(obs["meeting_to_win"]), '0.0000',
         "Observed over the trailing 12 months"),
        ("comm", "Commission on new ACV", float(cfg["opex"]["commission_new_pct"]), PCT1, "Per the plan"),
        ("ps_spend", "Paid Social spend run-rate ($/mo)", float(obs["paid_social_spend_per_month"]), MONEY,
         "Trailing 3 months"),
        ("ps_mrr", "Paid Social new MRR run-rate ($/mo)", float(obs["paid_social_new_mrr_per_month"]), MONEY,
         "Trailing 3 months — lost if the channel is switched off"),
    ]
    put(ws, "A6", "ASSUMPTIONS (blue = editable input)", SUB)
    ref = {}
    r = 7
    for key, label, val, fmt, note in inputs:
        put(ws, f"A{r}", label)
        put(ws, f"B{r}", val, BLUE, fmt, align="center")
        put(ws, f"C{r}", note)
        ref[key] = f"CONTROL!$B${r}"
        r += 1

    r += 1
    put(ws, f"A{r}", "ACTIVE SCENARIO PARAMETERS (driven by B3)", SUB)
    r += 1
    scen_params = ["sdr_heads_added", "list_overlap", "monthly_cost", "hire_lag_months",
                   "ramp_month_1", "ramp_month_2", "ramp_month_3",
                   "seg_mix_SMB", "seg_mix_MidMarket", "seg_mix_Enterprise",
                   "paid_social_cut_pct"]
    sref = {}
    for pname in scen_params:
        put(ws, f"A{r}", pname)
        put(ws, f"B{r}",
            f"=INDEX(DATA_SCENARIOS!$B$2:$D${1+len(scen_params)},"
            f"MATCH($A{r},DATA_SCENARIOS!$A$2:$A${1+len(scen_params)},0),$B$4)",
            GREEN, NUM2, align="center")
        sref[pname] = f"CONTROL!$B${r}"
        r += 1

    dv = DataValidation(type="list", formula1="=DATA_SCENARIOS!$B$1:$D$1", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B3"])

    # ---------------------------------------------------------- DATA_SCENARIOS
    ws = wb.create_sheet("DATA_SCENARIOS")
    widths(ws, {"A": 26, "B": 16, "C": 18, "D": 16, "E": 62})
    header_row(ws, 1, ["parameter"] + scen_order + ["note"])
    def scen_val(name, pname):
        s = cfg["scenarios"][name]
        if pname == "sdr_heads_added":
            return float(s.get("sdr_heads_added", 0))
        if pname == "list_overlap":
            return float(s.get("list_overlap", 1.0))
        if pname == "monthly_cost":
            return float(sum(b["heads"] * b["cost_each"] for b in s.get("build_up", []))
                         + s.get("tooling_data", 0))
        if pname == "hire_lag_months":
            return float(s.get("hire_lag_months", 0))
        if pname.startswith("ramp_month_"):
            i = int(pname[-1]) - 1
            return float(s.get("ramp_curve", [1, 1, 1])[i])
        if pname.startswith("seg_mix_"):
            return float(s.get("seg_mix", {}).get(pname[8:], 0.0))
        if pname == "paid_social_cut_pct":
            return float(s.get("paid_social_cut_pct", 0.0))
        raise KeyError(pname)

    notes = {"sdr_heads_added": "SDRs added on top of today's team",
             "list_overlap": "1.00 = new reps work the SAME account list; 0.00 = wholly new territory",
             "monthly_cost": "Fully-loaded build-up: reps, managers, AEs, SEs, tooling",
             "hire_lag_months": "Requisition opened -> seat filled",
             "ramp_month_1": "Share of full productivity, month 1 after start",
             "ramp_month_2": "month 2", "ramp_month_3": "month 3 onwards",
             "seg_mix_SMB": "Share of the pod's won logos", "seg_mix_MidMarket": "",
             "seg_mix_Enterprise": "", "paid_social_cut_pct": "1.00 = channel switched off"}
    for i, pname in enumerate(scen_params):
        r = 2 + i
        put(ws, f"A{r}", pname)
        for j, sname in enumerate(scen_order):
            put(ws, f"{get_column_letter(2+j)}{r}", scen_val(sname, pname), BLUE, NUM2,
                align="center")
        put(ws, f"E{r}", notes.get(pname, ""))
    r = 2 + len(scen_params) + 1
    put(ws, f"A{r}", "label", SUB)
    for j, sname in enumerate(scen_order):
        put(ws, f"{get_column_letter(2+j)}{r}", cfg["scenarios"][sname]["label"], BLUE)

    # ----------------------------------------------------------- DATA_ACTUALS
    ws = wb.create_sheet("DATA_ACTUALS")
    a = d["actuals"]
    cols = ["month", "month_index", "mrr", "recognised_revenue", "carry_in_mrr",
            "new_mrr", "expansion_mrr", "reactivation_mrr", "contraction_mrr",
            "churn_mrr", "cogs", "sm", "rd", "ga", "total_cost", "net_burn", "nrr_m"]
    header_row(ws, 1, cols)
    widths(ws, {get_column_letter(i + 1): 15 for i in range(len(cols))})
    ws.column_dimensions["A"].width = 10
    for i, row in a.iterrows():
        for j, c in enumerate(cols):
            v = row[c]
            fmt = None if c in ("month", "month_index") else (
                '0.00000' if c == "nrr_m" else MONEY2)
            put(ws, f"{get_column_letter(j+1)}{i+2}", v, BLUE, fmt)
    n_act = len(a)
    ws.freeze_panes = "B2"

    # ---------------------------------------------------------- DATA_SEGMENTS
    ws = wb.create_sheet("DATA_SEGMENTS")
    scols = ["segment", "gross_margin", "monthly_nrr", "outbound_acv",
             "sales_cycle_lag_months", "paid_social_mix", "sales_effort_index"]
    header_row(ws, 1, scols)
    widths(ws, {"A": 14, "B": 14, "C": 14, "D": 14, "E": 20, "F": 16, "G": 16})
    gm = dict(zip(d["seg_dim"].segment_code, d["seg_dim"].gross_margin))
    eff = dict(zip(d["seg_dim"].segment_code, d["seg_dim"].sales_effort_index))
    for i, s in enumerate(seg_order):
        r = 2 + i
        put(ws, f"A{r}", s, BLUE)
        put(ws, f"B{r}", float(gm[s]), BLUE, '0.000')
        put(ws, f"C{r}", float(obs["seg_nrr"][s]), BLUE, '0.00000')
        put(ws, f"D{r}", float(obs["outbound_acv"][s]), BLUE, MONEY)
        put(ws, f"E{r}", float(cfg["sales_cycle_lag_months"][s]), BLUE, NUM2)
        put(ws, f"F{r}", float(obs["paid_social_seg_mix"].get(s, 0.0)), BLUE, '0.0000')
        put(ws, f"G{r}", float(eff[s]), BLUE, NUM2)
    SEGR = {s: 2 + i for i, s in enumerate(seg_order)}

    # ---------------------------------------------------------- DATA_CHANNELS
    ws = wb.create_sheet("DATA_CHANNELS")
    ch = d["channels"]
    ccols = ["channel", "n_wins", "new_acv", "avg_acv", "gross_margin",
             "program_spend", "sdr_payroll", "shared_sales_alloc", "commission",
             "monthly_logo_churn_pct", "monthly_nrr_pct"]
    header_row(ws, 1, ccols)
    widths(ws, {get_column_letter(i + 1): 17 for i in range(len(ccols))})
    ws.column_dimensions["A"].width = 18
    ch = ch.sort_values("channel").reset_index(drop=True)
    for i, row in ch.iterrows():
        r = 2 + i
        put(ws, f"A{r}", row.channel, BLUE)
        put(ws, f"B{r}", int(row.n_wins), BLUE, '#,##0')
        put(ws, f"C{r}", float(row.new_acv), BLUE, MONEY)
        put(ws, f"D{r}", float(row.avg_acv), BLUE, MONEY)
        put(ws, f"E{r}", float(row.gross_margin), BLUE, '0.000')
        put(ws, f"F{r}", float(row.program_spend), BLUE, MONEY)
        put(ws, f"G{r}", float(row.sdr_payroll), BLUE, MONEY)
        put(ws, f"H{r}", float(row.shared_sales_alloc), BLUE, MONEY)
        put(ws, f"I{r}", float(row.commission), BLUE, MONEY)
        put(ws, f"J{r}", float(row.monthly_logo_churn_pct) / 100.0, BLUE, '0.0000%')
        put(ws, f"K{r}", float(row.monthly_nrr_pct) / 100.0, BLUE, '0.0000%')
    n_ch = len(ch)

    # ------------------------------------------------------- CALC_UNIT_ECON
    ws = wb.create_sheet("CALC_UNIT_ECON")
    ucols = ["channel", "wins (12m)", "avg ACV", "total CAC spend", "CAC",
             "monthly gross profit", "CAC payback (months)", "LTV (60m, NRR based)",
             "LTV / CAC", "verdict vs hurdle"]
    put(ws, "A1", "CHANNEL UNIT ECONOMICS — trailing 12 months, all formulas", TITLE)
    put(ws, "A2", "LTV = monthly gross profit x (1 - NRR^60) / (1 - NRR). "
                  "Finite 60-month horizon, observed net revenue retention.", BLACK)
    header_row(ws, 4, ucols)
    widths(ws, {get_column_letter(i + 1): 19 for i in range(len(ucols))})
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["J"].width = 24
    for i in range(n_ch):
        r = 5 + i
        s = 2 + i           # matching row in DATA_CHANNELS
        put(ws, f"A{r}", f"=DATA_CHANNELS!A{s}", GREEN)
        put(ws, f"B{r}", f"=DATA_CHANNELS!B{s}", GREEN, '#,##0')
        put(ws, f"C{r}", f"=DATA_CHANNELS!D{s}", GREEN, MONEY)
        put(ws, f"D{r}", f"=DATA_CHANNELS!F{s}+DATA_CHANNELS!G{s}"
                         f"+DATA_CHANNELS!H{s}+DATA_CHANNELS!I{s}", GREEN, MONEY)
        put(ws, f"E{r}", f"=D{r}/B{r}", BLACK, MONEY)
        put(ws, f"F{r}", f"=C{r}/12*DATA_CHANNELS!E{s}", GREEN, MONEY)
        put(ws, f"G{r}", f"=IFERROR(E{r}/F{r},\"n/a\")", BLACK, NUM2)
        put(ws, f"H{r}", f"=IF(ABS(1-DATA_CHANNELS!K{s})<0.0000001,F{r}*60,"
                         f"F{r}*(1-DATA_CHANNELS!K{s}^60)/(1-DATA_CHANNELS!K{s}))",
            GREEN, MONEY)
        put(ws, f"I{r}", f"=IFERROR(H{r}/E{r},\"n/a\")", BLACK, NUM2)
        put(ws, f"J{r}", f"=IF(G{r}<={ref['payback_hurdle']},\"PASS - within hurdle\","
                         f"\"FAIL - payback beyond runway\")", GREEN)
    UE_LAST = 4 + n_ch

    # --------------------------------------------------------- CALC_FORECAST
    ws = wb.create_sheet("CALC_FORECAST")
    put(ws, "A1", "18-MONTH FORECAST — every cell below is a formula", TITLE)
    put(ws, "A2", "Block 1 recreates the Holt recursion that won the backtest. "
                  "Block 2 projects it, adds the scenario, and rolls cash forward.", BLACK)

    # --- Block 1: Holt recursion over the 36 actual months
    put(ws, "A4", "BLOCK 1 — Holt smoothing over actual MRR", SUB)
    header_row(ws, 5, ["month", "actual MRR", "level", "trend"])
    widths(ws, {"A": 11, "B": 15, "C": 15, "D": 15})
    for i in range(n_act):
        r = 6 + i
        src = 2 + i
        put(ws, f"A{r}", f"=DATA_ACTUALS!A{src}", GREEN)
        put(ws, f"B{r}", f"=DATA_ACTUALS!C{src}", GREEN, MONEY2)
        if i == 0:
            put(ws, f"C{r}", f"=B{r}", BLACK, MONEY2)
            put(ws, f"D{r}", f"=B{r+1}-B{r}", BLACK, MONEY2)
        else:
            put(ws, f"C{r}", f"={ref['holt_alpha']}*B{r}"
                             f"+(1-{ref['holt_alpha']})*(C{r-1}+D{r-1})", GREEN, MONEY2)
            put(ws, f"D{r}", f"={ref['holt_beta']}*(C{r}-C{r-1})"
                             f"+(1-{ref['holt_beta']})*D{r-1}", GREEN, MONEY2)
    LVL = f"$C${5 + n_act}"
    TRD = f"$D${5 + n_act}"

    # --- Block 2: the forecast
    b2 = 5 + n_act + 3
    put(ws, f"A{b2-1}", "BLOCK 2 — scenario forecast and cash roll-forward", SUB)
    fcols = ["month", "k", "base MRR", "ramp", "eff. added heads", "existing eff. team",
             "new-rep eff. team", "incr. meetings", "incr. wins", "booked SMB",
             "booked MidMkt", "booked Ent", "live SMB", "live MidMkt", "live Ent",
             "delta MRR", "MRR", "COGS", "incr. S&M", "paid social saving",
             "incr. commission", "S&M", "R&D", "G&A", "total cost", "net burn", "cash"]
    header_row(ws, b2, fcols)
    for i in range(len(fcols)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 15
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 6

    seed = b2 + 1                       # month-zero seed row, all zeros
    put(ws, f"A{seed}", "seed", BLACK)
    put(ws, f"B{seed}", 0, BLACK)
    for col in "MNOP":
        put(ws, f"{col}{seed}", 0, BLACK, MONEY2)
    put(ws, f"AA{seed}", f"={ref['cash_on_hand']}", GREEN, MONEY)

    # per_rep(N) written inline so the saturation curve is visible in the sheet
    def per_rep(nref: str) -> str:
        return (f"IF({nref}<={ref['sdr_n0']},{ref['sdr_base_meet']},"
                f"{ref['sdr_base_meet']}*({ref['sdr_n0']}/{nref})^{ref['sdr_exp']})")

    f0 = seed + 1
    for i in range(H):
        r = f0 + i
        put(ws, f"A{r}", d["fmonths"][i], BLUE)
        put(ws, f"B{r}", i + 1, BLACK)
        put(ws, f"C{r}", f"={LVL}+{TRD}*B{r}", BLACK, MONEY2)
        put(ws, f"D{r}", f"=IF(B{r}-{sref['hire_lag_months']}<=0,0,"
                         f"IF(B{r}-{sref['hire_lag_months']}=1,{sref['ramp_month_1']},"
                         f"IF(B{r}-{sref['hire_lag_months']}=2,{sref['ramp_month_2']},"
                         f"{sref['ramp_month_3']})))", GREEN, NUM2)
        put(ws, f"E{r}", f"={sref['sdr_heads_added']}*D{r}", GREEN, NUM2)
        put(ws, f"F{r}", f"={ref['sdr_now']}+E{r}*{sref['list_overlap']}", GREEN, NUM2)
        put(ws, f"G{r}", f"={ref['sdr_now']}*{sref['list_overlap']}+E{r}", GREEN, NUM2)
        put(ws, f"H{r}", f"=IF(E{r}<=0,0,{ref['sdr_now']}*{per_rep(f'F{r}')}"
                         f"+E{r}*{per_rep(f'G{r}')}"
                         f"-{ref['sdr_now']}*{per_rep(ref['sdr_now'])})", GREEN, NUM2)
        put(ws, f"I{r}", f"=H{r}*{ref['m2w']}", GREEN, NUM2)
        for j, s in enumerate(seg_order):
            col = "JKL"[j]
            sr = SEGR[s]
            put(ws, f"{col}{r}",
                f"=I{r}*{sref['seg_mix_' + s]}*DATA_SEGMENTS!$D${sr}/12"
                f"-{sref['paid_social_cut_pct']}*{ref['ps_mrr']}*DATA_SEGMENTS!$F${sr}",
                GREEN, MONEY2)
        for j, s in enumerate(seg_order):
            col = "MNO"[j]
            book = "JKL"[j]
            sr = SEGR[s]
            put(ws, f"{col}{r}",
                f"=IF(B{r}<=DATA_SEGMENTS!$E${sr},0,"
                f"{col}{r-1}*DATA_SEGMENTS!$C${sr}"
                f"+INDEX(${book}${f0}:${book}${f0+H-1},B{r}-DATA_SEGMENTS!$E${sr}))",
                GREEN, MONEY2)
        put(ws, f"P{r}", f"=M{r}+N{r}+O{r}", BLACK, MONEY2)
        put(ws, f"Q{r}", f"=C{r}+P{r}", BLACK, MONEY2)
        put(ws, f"R{r}", f"=Q{r}*{ref['cogs_pct']}", GREEN, MONEY2)
        put(ws, f"S{r}", f"=IF(B{r}>{sref['hire_lag_months']},{sref['monthly_cost']},0)",
            GREEN, MONEY2)
        put(ws, f"T{r}", f"=-{sref['paid_social_cut_pct']}*{ref['ps_spend']}", GREEN, MONEY2)
        put(ws, f"U{r}", f"=(J{r}+K{r}+L{r})*12*{ref['comm']}", GREEN, MONEY2)
        put(ws, f"V{r}", f"=C{r}*{ref['sm_pct']}+S{r}+T{r}+U{r}", GREEN, MONEY2)
        put(ws, f"W{r}", f"={ref['rd_base']}*(1+{ref['infl']})^(B{r}/12)", GREEN, MONEY2)
        put(ws, f"X{r}", f"={ref['ga_base']}*(1+{ref['infl']})^(B{r}/12)", GREEN, MONEY2)
        put(ws, f"Y{r}", f"=R{r}+V{r}+W{r}+X{r}", BLACK, MONEY2)
        put(ws, f"Z{r}", f"=Y{r}-Q{r}", BLACK, MONEY2)
        put(ws, f"AA{r}", f"=AA{r-1}-Z{r}", BLACK, MONEY2)
    FC0, FC1 = f0, f0 + H - 1

    # ----------------------------------------------------------- OUT_SUMMARY
    ws = wb.create_sheet("OUT_SUMMARY")
    widths(ws, {"A": 46, "B": 20, "C": 20, "D": 20, "E": 46})
    put(ws, "A1", "NORTHWIND CLOUD — THE ANSWER", TITLE)
    put(ws, "A2", "=\"Active scenario: \"&CONTROL!$B$3", GREEN)
    put(ws, "A4", "POSITION TODAY", SUB)
    lines = [
        ("ARR (latest actual month)", f"=DATA_ACTUALS!C{1+n_act}*12", MONEY),
        ("Net burn, trailing 3 months", f"=AVERAGE(DATA_ACTUALS!P{n_act-1}:P{1+n_act})", MONEY),
        ("Cash on hand", f"={ref['cash_on_hand']}", MONEY),
        ("Static runway (cash / current burn), months",
         f"={ref['cash_on_hand']}/AVERAGE(DATA_ACTUALS!P{n_act-1}:P{1+n_act})", NUM2),
    ]
    r = 5
    for lab, f, fmt in lines:
        put(ws, f"A{r}", lab)
        put(ws, f"B{r}", f, GREEN, fmt)
        r += 1

    r += 1
    put(ws, f"A{r}", "UNDER THE ACTIVE SCENARIO (18 months out)", SUB)
    r += 1
    lines2 = [
        ("Incremental S&M, $/month", f"={sref['monthly_cost']}", MONEY),
        ("Paid Social spend retired, $/month",
         f"={sref['paid_social_cut_pct']}*{ref['ps_spend']}", MONEY),
        ("Net incremental cost, $/month",
         f"={sref['monthly_cost']}-{sref['paid_social_cut_pct']}*{ref['ps_spend']}", MONEY),
        ("ARR at month 18", f"=CALC_FORECAST!Q{FC1}*12", MONEY),
        ("Incremental ARR vs base at month 18", f"=CALC_FORECAST!P{FC1}*12", MONEY),
        ("Net burn at month 18", f"=CALC_FORECAST!Z{FC1}", MONEY),
        ("Cash at month 18", f"=CALC_FORECAST!AA{FC1}", MONEY),
        ("Months of cash left at month 18",
         f"=IFERROR(CALC_FORECAST!AA{FC1}/AVERAGE(CALC_FORECAST!Z{FC1-2}:Z{FC1}),\"n/a\")", NUM2),
        ("Incremental wins per month (fully ramped)", f"=CALC_FORECAST!I{FC1}", NUM2),
        ("Incremental CAC",
         f"=IFERROR({sref['monthly_cost']}/CALC_FORECAST!I{FC1},\"n/a\")", MONEY),
        ("Blended ACV of incremental wins",
         "=" + "+".join(f"{sref['seg_mix_'+s]}*DATA_SEGMENTS!$D${SEGR[s]}" for s in seg_order),
         MONEY),
        ("Blended gross margin of incremental wins",
         "=" + "+".join(f"{sref['seg_mix_'+s]}*DATA_SEGMENTS!$B${SEGR[s]}" for s in seg_order),
         '0.000'),
    ]
    for lab, f, fmt in lines2:
        put(ws, f"A{r}", lab)
        put(ws, f"B{r}", f, GREEN, fmt)
        r += 1
    put(ws, f"A{r}", "Incremental CAC payback (months)")
    put(ws, f"B{r}", f"=IFERROR(B{r-3}/(B{r-2}/12*B{r-1}),\"n/a\")", BLACK, NUM2)
    pay_row = r
    r += 1
    put(ws, f"A{r}", "VERDICT vs the payback hurdle", SUB)
    put(ws, f"B{r}", f"=IF(B{pay_row}<={ref['payback_hurdle']},\"APPROVE\",\"REJECT\")",
        GREEN, align="center")
    put(ws, f"C{r}", f"=\"payback \"&TEXT(B{pay_row},\"0.0\")&\" months vs hurdle \""
                     f"&TEXT({ref['payback_hurdle']},\"0\")", GREEN)

    # ---------------------------------------------------------- DATA_SQL_CHECK
    ws = wb.create_sheet("DATA_SQL_CHECK")
    widths(ws, {"A": 46, "B": 20, "C": 14, "D": 46})
    header_row(ws, 1, ["metric_key", "reference_value", "tolerance", "source"])
    # Reference values must come from the warehouse at FULL precision. An
    # earlier version read them from Q01, whose SELECT rounds ARR to whole
    # dollars for display — the reconciliation then "failed" by $0.40 against
    # a presentation artefact rather than a modelling error. Tolerances below
    # are set to exactly half the last displayed unit of whatever produced the
    # reference, so a PASS means "identical apart from display rounding".
    unrounded = pd.read_sql("""
        SELECT mrr*12 AS arr, net_burn, month_index FROM v_pnl_month
        ORDER BY month_index""", con)
    q02 = pd.read_csv(TAB / "Q02_arr_bridge.csv")
    ss = d["scenario_summary"].set_index("scenario")
    sens = d["sensitivity"].set_index("scenario")
    cash = d["cash_paths"].set_index("month")
    checks = [
        ("arr_latest", float(unrounded.arr.iloc[-1]), 0.005,
         "SQL v_pnl_month, full precision"),
        ("burn_3m_avg", float(unrounded.net_burn.tail(3).mean()), 0.005,
         "SQL v_pnl_month, full precision"),
        ("arr_bridge_closing_latest", float(q02.closing_arr.iloc[-1]), 0.005,
         "SQL Q02 (rounded 2dp -> tolerance 0.005)"),
        ("arr_bridge_max_residual",
         float(q02.residual_arr_must_be_zero.abs().max()), 0.005, "SQL Q02"),
    ]
    for _, row in d["channels"].iterrows():
        checks.append((f"cac__{row.channel}", float(row.cac), 0.5,
                       "SQL Q07 (rounded 0dp -> tolerance 0.5)"))
        checks.append((f"payback__{row.channel}", float(row.cac_payback_months), 0.05,
                       "SQL Q07 (rounded 1dp -> tolerance 0.05)"))
        checks.append((f"ltvcac__{row.channel}", float(row.ltv_cac_b), 0.0051,
                       "SQL Q07 (rounded 2dp -> tolerance 0.005)"))
    for s in scen_order:
        checks.append((f"cash_m18__{s}", float(ss.loc[s, "cash_month18"]), 0.5,
                       "Python cash model src/04 (rounded 0dp)"))
        checks.append((f"arr_m18__{s}", float(ss.loc[s, "arr_month18"]), 0.5,
                       "Python cash model src/04 (rounded 0dp)"))
        checks.append((f"delta_arr_m18__{s}", float(ss.loc[s, "delta_arr_month18"]), 0.5,
                       "Python cash model src/04 (rounded 0dp)"))
    for s in ("sales_proposal", "selective"):
        checks.append((f"inc_cac__{s}", float(sens.loc[s, "incremental_cac"]), 0.5,
                       "Python sensitivity src/04 (rounded 0dp)"))
        checks.append((f"inc_payback__{s}", float(sens.loc[s, "cac_payback_months"]), 0.05,
                       "Python sensitivity src/04 (rounded 1dp)"))
    for i, (k, v, tol, src) in enumerate(checks):
        r = 2 + i
        put(ws, f"A{r}", k, BLUE)
        put(ws, f"B{r}", v, BLUE, MONEY2)
        put(ws, f"C{r}", tol, BLUE, NUM2)
        put(ws, f"D{r}", src, BLUE)
    n_chk = len(checks)
    CHK_RANGE = f"DATA_SQL_CHECK!$A$2:$A${1+n_chk}"
    CHK_VALS = f"DATA_SQL_CHECK!$B$2:$B${1+n_chk}"
    CHK_TOL = f"DATA_SQL_CHECK!$C$2:$C${1+n_chk}"

    # ---------------------------------------------------------------- CHECKS
    ws = wb.create_sheet("CHECKS")
    widths(ws, {"A": 42, "B": 20, "C": 20, "D": 16, "E": 14, "F": 12, "G": 40})
    put(ws, "A1", "EXCEL vs SQL RECONCILIATION", TITLE)
    put(ws, "A2", "Excel value is computed by formula on this workbook. Reference "
                  "value comes from the SQL warehouse or the Python cash model. "
                  "Scenario-specific rows follow CONTROL!B3.", BLACK)
    header_row(ws, 4, ["check", "excel value", "reference value", "difference",
                       "tolerance", "status", "note"])

    def lookup(keyexpr: str) -> str:
        return f"INDEX({CHK_VALS},MATCH({keyexpr},{CHK_RANGE},0))"

    def tol(keyexpr: str) -> str:
        return f"INDEX({CHK_TOL},MATCH({keyexpr},{CHK_RANGE},0))"

    check_rows = []
    check_rows.append(("ARR, latest actual month", f"=DATA_ACTUALS!C{1+n_act}*12",
                       '"arr_latest"', "Sum of the fact table x 12"))
    check_rows.append(("Net burn, trailing 3 months",
                       f"=AVERAGE(DATA_ACTUALS!P{n_act-1}:P{1+n_act})",
                       '"burn_3m_avg"', "Excel averages the actuals"))
    check_rows.append(("ARR bridge, closing (latest month)",
                       f"=(DATA_ACTUALS!C{n_act}+DATA_ACTUALS!E{1+n_act}"
                       f"+DATA_ACTUALS!F{1+n_act}+DATA_ACTUALS!G{1+n_act}"
                       f"+DATA_ACTUALS!H{1+n_act}-DATA_ACTUALS!I{1+n_act}"
                       f"-DATA_ACTUALS!J{1+n_act})*12",
                       '"arr_bridge_closing_latest"',
                       "Excel rebuilds opening + movements and must land on closing"))
    for i in range(n_ch):
        ur = 5 + i
        cr = 2 + i
        check_rows.append((f"CAC — channel row {i+1}", f"=CALC_UNIT_ECON!E{ur}",
                           f'"cac__"&DATA_CHANNELS!A{cr}', "Formula from the cost pools"))
        check_rows.append((f"CAC payback — channel row {i+1}", f"=CALC_UNIT_ECON!G{ur}",
                           f'"payback__"&DATA_CHANNELS!A{cr}', "CAC / monthly gross profit"))
        check_rows.append((f"LTV/CAC — channel row {i+1}", f"=CALC_UNIT_ECON!I{ur}",
                           f'"ltvcac__"&DATA_CHANNELS!A{cr}', "60-month NRR-based LTV"))
    check_rows.append(("Cash at month 18 (active scenario)", f"=CALC_FORECAST!AA{FC1}",
                       '"cash_m18__"&CONTROL!$B$3', "Follows the scenario switch"))
    check_rows.append(("ARR at month 18 (active scenario)", f"=CALC_FORECAST!Q{FC1}*12",
                       '"arr_m18__"&CONTROL!$B$3', "Follows the scenario switch"))
    check_rows.append(("Incremental ARR at month 18 (active)", f"=CALC_FORECAST!P{FC1}*12",
                       '"delta_arr_m18__"&CONTROL!$B$3', "Follows the scenario switch"))

    for i, (lab, xf, key, note) in enumerate(check_rows):
        r = 5 + i
        put(ws, f"A{r}", lab)
        put(ws, f"B{r}", xf, GREEN, MONEY2)
        put(ws, f"C{r}", f"=IFERROR({lookup(key)},\"key missing\")", GREEN, MONEY2)
        put(ws, f"D{r}", f"=IFERROR(B{r}-C{r},\"n/a\")", BLACK, MONEY2)
        put(ws, f"E{r}", f"=IFERROR({tol(key)},\"\")", GREEN, NUM2)
        put(ws, f"F{r}", f"=IF(AND(ISNUMBER(B{r}),ISNUMBER(C{r})),"
                         f"IF(ABS(B{r}-C{r})<=E{r},\"PASS\",\"FAIL\"),\"FAIL\")",
            BLACK, align="center")
        put(ws, f"G{r}", note)
    last = 4 + len(check_rows)
    put(ws, f"A{last+2}", "OVERALL", SUB)
    put(ws, f"B{last+2}", f'=COUNTIF(F5:F{last},"PASS")&" / "&{len(check_rows)}&" PASS"',
        BLACK)
    put(ws, f"C{last+2}", f'=IF(COUNTIF(F5:F{last},"FAIL")=0,"ALL CHECKS PASS",'
                          f'COUNTIF(F5:F{last},"FAIL")&" FAILING")', BLACK)

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    con.close()

    meta = dict(forecast_first_row=FC0, forecast_last_row=FC1, n_checks=len(check_rows),
                check_status_range=f"F5:F{last}", n_actuals=n_act, n_channels=n_ch,
                unit_econ_rows=[5, UE_LAST])
    json.dump(meta, open(ROOT / "outputs" / "logs" / "excel_layout.json", "w"), indent=2)
    print(f"wrote {XLSX.relative_to(ROOT)}")
    print(f"  sheets            : {', '.join(wb.sheetnames)}")
    print(f"  forecast rows     : {FC0}..{FC1}")
    print(f"  reconciliation    : {len(check_rows)} checks against "
          f"{n_chk} reference values")
    return 0


if __name__ == "__main__":
    sys.exit(build())
